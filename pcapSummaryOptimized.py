
import pandas as pd
from pandas import *
import xlsxwriter
from os import listdir
import collections
from time import sleep
import os
from os import path
from datetime import datetime
import openpyxl

#import numpy as np
#Global variables
row                 =  0
col                 =  0
pcapSummaryList     =  []
pcapSummary         =  collections.OrderedDict()


#Method to get List of pcap files in current directory
def getAllPcapFiles():
	pcapFileList = []
	for f in listdir(os.getcwd()):
		if f.endswith('.' + "pcap"):
			pcapFileList.append(f)
	
	return pcapFileList
def initialize():
	#  pcapSummary = collections.OrderedDict()
	pcapSummary['pcapFileName']                =      ""
	pcapSummary['totalPackets']                =      0
	pcapSummary['totalManagementFrames']       =      0
	pcapSummary['totalControlFrames']          =      0
	pcapSummary['totalDataFrames']             =      0
	pcapSummary['numberOfRetryULFrames']       =      0
	pcapSummary['numberOfRetryDLFrames']       =      0
	pcapSummary['totalDataULFrames']           =      0
	pcapSummary['totalDataDLFrames']           =      0
	pcapSummary['avgDataDLLength']             =      0
	pcapSummary['avgDataULLength']             =      0
	pcapSummary['avgChanUtilizationLength']    =      0
	pcapSummary['avgAirTimeUtilization']       =      0
		
	for i in range(16):
		UL='UL'+`i`
		pcapSummary[UL]=0
	for i in range(16):
		DL='DL'+`i`
		pcapSummary[DL]=0
	pcapSummary['AverageULMCS']                =      0
	pcapSummary['AverageDLMCS']                =      0
################Method to get AirUtilization###########
def AirtimeUtilization(packet,duration):
	ST                          =    9
	SIFS                        =    16
	QOSclass                    =    0
	sumOFAirU                   =    0
	BackOff                     =    0
	sum                         =    0
	time                        =    os.popen('capinfos  -T -r -a -e '+packet+'').read().split("\n")
	QOSCLASS                    =    os.popen('tshark -r '+packet+'  -Y "(radiotap.dbm_antsignal >-82) && (wlan.qos.priority) && (wlan.fcs.status=="Good"") -T fields -e frame.number -e  wlan.qos.priority -e wlan_radio.duration ' ).read().split("\n")

	time                        =    time[0].split("\t")
	time_1                      =    time[1]
	time_2                      =    time[2]
	start                       =    datetime.strptime(time_1, "%Y-%m-%d %H:%M:%S.%f")
	end                         =    datetime.strptime(time_2, "%Y-%m-%d %H:%M:%S.%f")
	diff_time                   =    end - start
	diff_time                   =    diff_time.seconds
	
	
	
	for i in range(len(QOSCLASS)-1):
		QOSCLASS[i]=QOSCLASS[i].split("\t")
	for i in range(len(QOSCLASS)-1):
		QOS=int(QOSCLASS[i][1])
		#Duration=int(cmdQOSCLASS[i][2])
		Duration=0
		
		if QOS == 0 or QOS == 3:
			AIFS=QOS*ST+SIFS
			#BackOff=15+((1023-15)/2)
			AirUtilization=(Duration+AIFS+BackOff)
		if QOS == 1 or QOS == 2:
			AIFS=QOS*ST+SIFS
			#BackOff=15+((1023-15)/2)
			AirUtilization=(Duration+AIFS+BackOff)
		if QOS== 4 or QOS == 5:
			AIFS=QOS*ST+SIFS
			#BackOff=7+((15-7)/2)
			AirUtilization=(Duration+AIFS+BackOff)
		if QOS == 6 or QOS == 7:
			AIFS=QOS*ST+SIFS
			#BackOff=3+((7-3)/2)
			AirUtilization=(Duration+AIFS+BackOff)

		sumOFAirU+=AirUtilization


	sumOFAirU =sumOFAirU+duration
	AvgOFAirU  =  ((float(sumOFAirU))/(diff_time*1000000))*100
	pcapSummary['avgAirTimeUtilization'] =  round(AvgOFAirU,2)





################Method to get PcapSummary#############
def getPcapSummary(packet):
	sumofCU   =          0
	mgmt      =          0
	ctrl      =          0
	data      =          0
	duration  =          0
	DlMCS     =          []
	DlRetry   =          []
	DlFlen    =          0
	DLFrame   =          0
	UlMCS     =          []
	UlRetry   =          []
	UlFlen    =          0
	ULFrame   =          0
	framelen  =          0
	#We are finding the type of frame and total duration
	#From type extract the management control and data frame
	type      =          os.popen('tshark -r '+packet+'   -I -T fields -e wlan.fc.type -e wlan_radio.duration -e frame.len').read().split("\n")
	for pac in range(len(type)-1):
		type[pac]=type[pac].split("\t")
	#Extract the mcs index, retry, frame length
	subtype   =          os.popen('tshark -r '+packet+'  -I  -Y  "wlan.fc.ds==1 || wlan.fc.ds==2"   -T fields -e wlan.fc.ds -e  radiotap.mcs.index -e wlan.fc.retry -e frame.len').read().split("\n")
	for pac in range(len(subtype)-1):
		subtype[pac]=subtype[pac].split("\t")
	#Extracts the Channel Utilization
	CU        =          os.popen('tshark -r '+packet+'  -I -Y  "wlan.qbss.cu"  -T fields -e wlan.qbss.cu ').read().split("\n")
	for bytes in range(len(CU)-1):
		sumofCU=sumofCU+int(CU[bytes])
	
	

	#Calculating the number of Managemnt ,Control, Data frame,and total Frame Length
	for pac in range(len(type)-1):
		if type[pac][0]!='' and len(type[pac][0])<2:
			if int(type[pac][0])==0:
				mgmt+=1
			if int(type[pac][0])==1:
				ctrl+=1
			if int(type[pac][0])==2:
				data+=1
		if type[pac][1]!='' :
			duration+=int(type[pac][1])
		if type[pac][2]!='':
			framelen+=int(type[pac][2])
	#Calculating the total UL,DL ,Retry in UL,Retry in DL,MCS in UL,MCS in DL,total UL Frame Length ,total DL Frame Length
	for pac in range(len(subtype)-1):
		if subtype[pac][0]!='' :
			sub = subtype[pac][0].split("x")
			if len(sub[1])<9:
				if int(sub[1]) == 1:
					ULFrame=ULFrame+1
					if subtype[pac][1]!='' :
						UlMCS.append(int(subtype[pac][1]))
					if subtype[pac][2]!='' :
						UlRetry.append(int(subtype[pac][2]))
					if subtype[pac][3]!='' :
						UlFlen=UlFlen+(int(subtype[pac][3]))
				if int(sub[1]) == 2:
					DLFrame=DLFrame+1
					if subtype[pac][1]!='' :
						DlMCS.append(int(subtype[pac][1]))
					if subtype[pac][2]!='' :
						DlRetry.append(int(subtype[pac][2]))
					if subtype[pac][3]!='' :
						DlFlen=DlFlen+(int(subtype[pac][3]))
		

	#Assigning the value to the specific key
	pcapSummary['totalPackets']                 =   len(type)
	pcapSummary['totalManagementFrames']        =   mgmt
	pcapSummary['totalControlFrames']           =   ctrl
	pcapSummary['totalDataFrames']              =   data
	pcapSummary['numberOfRetryULFrames']        =   len(UlRetry)
	pcapSummary['numberOfRetryDLFrames']        =   len(DlRetry)
	pcapSummary['totalDataULFrames']            =   ULFrame
	pcapSummary['totalDataDLFrames']            =   DLFrame
	pcapSummary['avgDataDLLength']              =   DlFlen/DLFrame
	pcapSummary['avgDataULLength']              =   UlFlen/ULFrame
	if  sumofCU!=0 :
		pcapSummary['avgChanUtilizationLength'] =   sumofCU/len(CU)
	#Total Air utilization while traffic is running
	AirtimeUtilization(packet,duration)
	#Finding the particular MCS index and how many frames comes in specific index
	getMCSDistribution(packet,UlMCS,DlMCS)

######################Method to Get MCS Distribution##############################
def getMCSDistribution(packet,UlMCS,DlMCS):
	a                                           =    [None]*17
	b                                           =    [None]*17
	PerMCSUL                                    =     0
	TotalMCSUL                                  =     0
	TotalULPacket                               =     0
	PerMCSDL                                    =     0
	TotalMCSDL                                  =     0
	TotalDLPacket                               =     0
	for i in range(16):
		a[i]                                    =     0
		b[i]                                    =     0
	#Calculating the Per MCSindex how many frames asides for UL
	for index in range(len(UlMCS)):
		if UlMCS[index]<16:
			a[UlMCS[index]]+=1
	for index in range (len(a)-1):
		PerMCSUL = a[index]*index
		TotalMCSUL+=PerMCSUL
		TotalULPacket+=a[index]
	#Calculating the Per MCSindex how many frames asides for DL
	for index in range(len(DlMCS)):
		if DlMCS[index]<16:
			b[DlMCS[index]]+=1
	for index in range (len(b)-1):
		PerMCSDL = b[index]*index
		TotalMCSDL+=PerMCSDL
		TotalDLPacket+=b[index]
	#Assigning the number of packets of per MCS index
	for index in range(16):
		UL='UL'+`index`
		pcapSummary[UL]=a[index]
	for index in range(16):
		DL='DL'+`index`
		pcapSummary[DL]=b[index]
	#Finding the Average of total MCS index in UL and DL
	if TotalMCSUL!=0 or TotalULPacket!=0 :
		pcapSummary['AverageULMCS']=TotalMCSUL/TotalULPacket
	if  TotalMCSDL!=0 or TotalDLPacket!=0:
		pcapSummary['AverageDLMCS']=TotalMCSDL/TotalDLPacket



def update(xlsxFileName,row):
	print os.getcwd()
	dir=os.chdir('../Analytics')
	wb = openpyxl.load_workbook(xlsxFileName+'.xlsx')
	worksheet =wb.active
	for eachPcapSummary in pcapSummaryList:
		worksheet.append(eachPcapSummary.values())
		break
		
		
	wb.save(xlsxFileName+'.xlsx')

#####################################################################################
def writePcapSummaryToXlsx(xlsxFileName):

	
	global row
	global col
	dir=os.chdir('../Analytics')
	workbook= xlsxwriter.Workbook(os.getcwd()+"/"+xlsxFileName+'.xlsx')
	worksheet=workbook.add_worksheet()
	format = workbook.add_format({'bold': True, 'font_color': 'black'})
	#Create header for xlsx sheet
	for key_header in pcapSummary.keys():
		worksheet.write(row,col,key_header,format)
		col += 1
	row += 1
	
	#Create pcapSummaryList each dict values to xlsx sheet
	for eachPcapSummary in pcapSummaryList:
		col = 0
		print (eachPcapSummary.values())
		for value_summary in eachPcapSummary.values():
			worksheet.write(row,col,value_summary)
			col += 1
		row += 1
	xlsxwriter.Workbook(os.getcwd()+"/"+xlsxFileName+'.xlsx').close()
	


###################Main function#######################################################
def main():
	run=0
	pcapFileList = getAllPcapFiles()
	xlsxFileName=raw_input("Enter the excel sheet name according to the AP: ")
	for eachPcap in pcapFileList:
		run=0
		initialize()
		dir=os.chdir('../Analytics')
		exists=(str(path.exists(xlsxFileName+'.xlsx')))
		#print exists
		if exists=="True":
			df = pd.read_excel(xlsxFileName+'.xlsx')
			#print df['pcapFileName']
			length=len(df['pcapFileName'])
			for i in range(0,length):
				print eachPcap.strip('.pcap')
				if eachPcap.strip('.pcap')!=df['pcapFileName'][i]:
					print "Not Present"
					run=1
					
				else:
					run=0
					print "Present"
					break
					

			if run==1:
				os.chdir('../test')
				pcapSummary['pcapFileName'] = eachPcap.strip('.pcap')
				print "Checking "+os.getcwd()+" "+eachPcap+" file"
				getPcapSummary(os.getcwd()+"/"+eachPcap)
				pcapSummaryTemp = pcapSummary
				pcapSummaryList.append(pcapSummaryTemp)
				update(xlsxFileName,length)
				print pcapSummary
				

		else:
			os.chdir('../test')
			pcapSummary['pcapFileName'] = eachPcap.strip('.pcap')
		#print the cureent working directory and the current pcap file
			print "Checking "+os.getcwd()+" "+eachPcap+"file"
			getPcapSummary(os.getcwd()+"/"+eachPcap)
			pcapSummaryTemp = pcapSummary
			pcapSummaryList.append(pcapSummaryTemp)
			#pcapSummaryList[0]=pcapSummaryTemp
			print pcapSummary
			writePcapSummaryToXlsx(xlsxFileName)
		#print the cureent working directory and the current pcap file
##################################################################################

if __name__ == '__main__':
	main()`